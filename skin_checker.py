import os
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# Disable SSL warnings since League's local API uses a self-signed certificate
requests.packages.urllib3.disable_warnings()

def get_lockfile_data():
    """Finds the League client lockfile and extracts port and password."""
    possible_paths = [
        "D:/P1/Riot Games/League of Legends/lockfile",
        "C:/Riot Games/League of Legends/lockfile",
        "D:/Riot Games/League of Legends/lockfile"
    ]
    
    lockfile_path = None
    for path in possible_paths:
        if os.path.exists(path):
            lockfile_path = path
            break
            
    if not lockfile_path:
        raise FileNotFoundError("Could not find your League lockfile. Is the game open?")
        
    with open(lockfile_path, 'r') as f:
        content = f.read()
        
    parts = content.split(':')
    return {"port": parts[2], "password": parts[3]}

def get_official_skin_metadata():
    """Fetches up-to-date skin rarity data mapping from CommunityDragon."""
    print("Downloading skin rarity matrix...")
    try:
        url = "https://raw.communitydragon.org/latest/plugins/rcp-be-lol-game-data/global/default/v1/skins.json"
        res = requests.get(url)
        if res.status_code == 200:
            raw_skins = res.json()
            return {int(skin_id): data for skin_id, data in raw_skins.items()}
    except Exception as e:
        print(f"Warning: Could not fetch rarity metadata. Error: {e}")
    return {}

def fetch_skins_data():
    try:
        lockfile = get_lockfile_data()
    except Exception as e:
        print(f"Error: {e}")
        return

    meta_dict = get_official_skin_metadata()
    auth = ('riot', lockfile['password'])
    base_url = f"https://127.0.0.1:{lockfile['port']}"
    
    print("Connecting to LCU API...")
    
    # 1. Fetch Dynamic Champion Manifest
    champ_map_url = f"{base_url}/lol-champions/v1/owned-champions-minimal"
    champ_res = requests.get(champ_map_url, auth=auth, verify=False)
    
    champ_dict = {}
    all_champions_list = []
    champion_counts_tracker = {}
    
    if champ_res.status_code == 200:
        for c in champ_res.json():
            champ_dict[int(c['id'])] = c['name']
            all_champions_list.append(c['name'])
            champion_counts_tracker[c['name']] = 0
    
    all_champions_list.sort()
            
    # Fetch Summoner ID
    summoner_url = f"{base_url}/lol-summoner/v1/current-summoner"
    sum_res = requests.get(summoner_url, auth=auth, verify=False)
    if sum_res.status_code != 200:
        print("Could not fetch current summoner profile.")
        return
    summoner_id = sum_res.json().get('summonerId')
    
    # Query client skin inventory
    skins_url = f"{base_url}/lol-champions/v1/inventories/{summoner_id}/skins-minimal"
    print("Compiling inventory ledger...")
    response = requests.get(skins_url, auth=auth, verify=False)
    
    if response.status_code != 200:
        print("Failed to fetch skins.")
        return
        
    skins_data = response.json()
    owned_skin_ids = set()
    owned_skins = []
    
    for item in skins_data:
        if item.get('ownership', {}).get('owned') is True and not item.get('isBase'):
            skin_id = int(item.get('id', 0))
            owned_skin_ids.add(skin_id)
            skin_name = item.get('name', 'Unknown Skin')
            rarity_tier = "Budget / Legacy / Promo"
            
            champ_id_key = skin_id // 1000
            champion_name = champ_dict.get(champ_id_key, "Unknown Champion")
            
            if champion_name in champion_counts_tracker:
                champion_counts_tracker[champion_name] += 1
            
            if skin_id in meta_dict:
                meta = meta_dict[skin_id]
                rarity_item = meta.get('rarity', '')
                
                if "Epic" in rarity_item or rarity_item == "kEpic":
                    rarity_tier = "Epic (1350 RP)"
                elif "Legendary" in rarity_item or rarity_item == "kLegendary":
                    rarity_tier = "Legendary (1820 RP)"
                elif "Mythic" in rarity_item or rarity_item == "kMythic":
                    rarity_tier = "Mythic / Hextech"
                elif "Ultimate" in rarity_item or rarity_item == "kUltimate":
                    rarity_tier = "Ultimate (3250 RP)"
            
            owned_skins.append({
                "Champion": champion_name,
                "Skin Name": skin_name,
                "Rarity": rarity_tier,
                "Item ID": str(skin_id)
            })
                    
    owned_skins.sort(key=lambda x: x['Champion'])
    
    # Calculate pool capacities per champion
    rollable_remaining = {name: 0 for name in all_champions_list}
    unrollable_remaining = {name: 0 for name in all_champions_list}
    
    for skin_id, data in meta_dict.items():
        if skin_id % 1000 == 0:
            continue
            
        champ_id = skin_id // 1000
        champ_name = champ_dict.get(champ_id)
        
        if champ_name in rollable_remaining:
            if skin_id in owned_skin_ids:
                continue
                
            skin_name = data.get('name', '')
            if "Chroma" in skin_name or "Placeholder" in skin_name:
                continue
                
            rarity = data.get('rarity', '')
            if "Limited" in rarity or "kNoRarity" in rarity or data.get('isLegacy') == "Limited":
                unrollable_remaining[champ_name] += 1
            else:
                rollable_remaining[champ_name] += 1

    # Calculate max limit ceiling
    max_skins_cap = max(champion_counts_tracker.values()) if champion_counts_tracker else 0
    
    # Group champions by owned count for the distribution view layout
    distribution_map = {i: [] for i in range(max_skins_cap + 1)}
    for champ, count in champion_counts_tracker.items():
        if count <= max_skins_cap:
            distribution_map[count].append(champ)
            
    for count in distribution_map:
        distribution_map[count].sort()
    
    # ==================== EXCEL CONFIGURATION ====================
    print("Building progress analytics spreadsheet...")
    wb = openpyxl.Workbook()
    
    font_name = "Segoe UI"
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid") # Charcoal
    sub_header_fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid") # Secondary Gray
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid") # Off-white
    
    header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_name, size=10)
    border_side = Side(border_style="thin", color="E5E7EB")
    grid_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Custom format masks for formatting numbers as professional hyphens
    dash_number_format = '#,##0;-#,##0;"-"'
    
    # --- TAB 1: Champion Skin Summary ---
    ws1 = wb.active
    ws1.title = "Champion Skin Summary"
    ws1.views.sheetView[0].showGridLines = False
    
    # Left Dashboard Panel Headers (Columns A through E)
    headers_ws1_left = [
        "Champion", "Skins Owned", "Rollable Skins Remaining", 
        "Unrollable Skins Remaining", "% of Skins Owned"
    ]
    for col_num, header in enumerate(headers_ws1_left, 1):
        cell = ws1.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col_num > 1 else "left")
        
    # Build Dashboard Metrics Data rows
    for row_num, champ_name in enumerate(all_champions_list, 2):
        cell_name = ws1.cell(row=row_num, column=1, value=champ_name)
        cell_name.font = data_font
        cell_name.border = grid_border
        
        formula_owned = f"=COUNTIF('Owned Skin Details'!A:A, \"{champ_name}\")"
        cell_owned = ws1.cell(row=row_num, column=2, value=formula_owned)
        cell_owned.font = data_font
        cell_owned.border = grid_border
        cell_owned.number_format = dash_number_format
        cell_owned.alignment = Alignment(horizontal="center")
        
        cell_rollable = ws1.cell(row=row_num, column=3, value=rollable_remaining[champ_name])
        cell_rollable.font = data_font
        cell_rollable.border = grid_border
        cell_rollable.number_format = dash_number_format
        cell_rollable.alignment = Alignment(horizontal="center")
        
        cell_unrollable = ws1.cell(row=row_num, column=4, value=unrollable_remaining[champ_name])
        cell_unrollable.font = data_font
        cell_unrollable.border = grid_border
        cell_unrollable.number_format = dash_number_format
        cell_unrollable.alignment = Alignment(horizontal="center")
        
        # Dynamic Completion Ratio Formula tracking balance logic: Owned / (Owned + Rollable + Unrollable)
        formula_pct = f"=IFERROR(B{row_num} / (B{row_num} + C{row_num} + D{row_num}), 0)"
        cell_pct = ws1.cell(row=row_num, column=5, value=formula_pct)
        cell_pct.font = data_font
        cell_pct.border = grid_border
        # CHANGED: Formatted as whole number percentage with a hyphen fallback for 0%
        cell_pct.number_format = '0%;-0%;"-"'
        cell_pct.alignment = Alignment(horizontal="center")
        
        if row_num % 2 == 0:
            for c in [cell_name, cell_owned, cell_rollable, cell_unrollable, cell_pct]:
                c.fill = zebra_fill

    # Right Distribution Panels (Columns G, H, I)
    headers_ws1_right = ["Number of Skins", "Count of Champions", "Champions"]
    for idx, header in enumerate(headers_ws1_right):
        col_num = 7 + idx # 7 = G, 8 = H, 9 = I
        cell = ws1.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = sub_header_fill
        cell.alignment = Alignment(horizontal="left" if col_num == 9 else "center")
        
    # Generate distribution summaries
    for dist_idx, bracket_size in enumerate(range(max_skins_cap + 1), 2):
        cell_bracket = ws1.cell(row=dist_idx, column=7, value=f"{bracket_size} Skins")
        cell_bracket.font = data_font
        cell_bracket.border = grid_border
        cell_bracket.alignment = Alignment(horizontal="left")
        
        dist_formula = f"=COUNTIF(B:B, {bracket_size})"
        cell_dist_count = ws1.cell(row=dist_idx, column=8, value=dist_formula)
        cell_dist_count.font = data_font
        cell_dist_count.border = grid_border
        cell_dist_count.number_format = dash_number_format
        cell_dist_count.alignment = Alignment(horizontal="center")
        
        champs_in_bracket = distribution_map.get(bracket_size, [])
        champs_string = ", ".join(champs_in_bracket)
        
        cell_list = ws1.cell(row=dist_idx, column=9, value=champs_string)
        cell_list.font = data_font
        cell_list.border = grid_border
        cell_list.alignment = Alignment(horizontal="left", wrap_text=False)
        
        if dist_idx % 2 == 0:
            cell_bracket.fill = zebra_fill
            cell_dist_count.fill = zebra_fill
            cell_list.fill = zebra_fill

    # Set Spacer Column F to exactly 21 width units (~150 pixels)
    ws1.column_dimensions['F'].width = 21

    # --- TAB 2: Owned Skin Details ---
    ws2 = wb.create_sheet(title="Owned Skin Details")
    ws2.views.sheetView[0].showGridLines = False
    
    headers_ws2 = ["Champion", "Skin Name", "Rarity", "Item ID"]
    for col_num, header in enumerate(headers_ws2, 1):
        cell = ws2.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col_num > 2 else "left")
        
    for row_num, skin in enumerate(owned_skins, 2):
        row_values = [skin["Champion"], skin["Skin Name"], skin["Rarity"], skin["Item ID"]]
        for col_num, val in enumerate(row_values, 1):
            cell = ws2.cell(row=row_num, column=col_num, value=val)
            cell.font = data_font
            cell.border = grid_border
            if row_num % 2 == 0:
                cell.fill = zebra_fill
            if col_num == 4:
                cell.number_format = "@"
                cell.alignment = Alignment(horizontal="left")

    # Column width layout adjuster engine
    for ws in [ws1, ws2]:
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            
            if ws == ws1 and col_letter == 'F':
                continue
                
            if ws == ws1 and col_letter == 'I':
                ws.column_dimensions[col_letter].width = 30
                continue
                
            max_len = 0
            for cell in col:
                val_to_check = str(cell.value or '')
                if cell.value and str(cell.value).startswith('='):
                    val_to_check = "12"
                if len(val_to_check) > max_len:
                    max_len = len(val_to_check)
            ws.column_dimensions[col_letter].width = max(max_len + 5, 14)
            
    # Conditional formatting color rule for 0 skins alerts on Tab 1 Column B
    zero_fill = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")
    zero_font = Font(name=font_name, size=10, color="EF4444", bold=True)
    ws1.conditional_formatting.add(
        f"B2:B{len(all_champions_list)+1}",
        CellIsRule(operator='equal', formula=['0'], stopIfTrue=True, fill=zero_fill, font=zero_font)
    )

    output_file = "my_league_skins.xlsx"
    wb.save(output_file)
    print(f"\nSuccess! Percentages updated to whole numbers.")
    print(f"File updated here: {os.path.abspath(output_file)}")

if __name__ == "__main__":
    fetch_skins_data()